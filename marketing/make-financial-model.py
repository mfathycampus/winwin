# -*- coding: utf-8 -*-
"""WINوWIN — Detailed financial model (5 years, 3 scenarios). Arabic RTL, live formulas."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
import os

NAVY = '0E2254'; GREEN = '1B8A4A'; GOLD = 'B8860B'
BLUE = '0000FF'; GREENTXT = '008000'; BLACK = '000000'
HEADFILL = PatternFill('solid', fgColor=NAVY)
SUBFILL = PatternFill('solid', fgColor='EEF4F1')
YELLOW = PatternFill('solid', fgColor='FFF6D6')
TOTALFILL = PatternFill('solid', fgColor='E9EEF5')
thin = Side(style='thin', color='B4C4BC')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
F = 'Arial'
MONEY = '#,##0;(#,##0);-'
PCT = '0.0%;(0.0%);-'
MULT = '0.0x'

wb = openpyxl.Workbook()

def style(ws):
    ws.sheet_view.rightToLeft = True

def label(ws, cell, text, bold=False, color=BLACK, size=11, fill=None, align='right'):
    c = ws[cell]; c.value = text
    c.font = Font(name=F, bold=bold, color=color, size=size)
    c.alignment = Alignment(horizontal=align, vertical='center', readingOrder=2, wrap_text=False)
    if fill: c.fill = fill
    return c

def num(ws, cell, value, fmt=MONEY, color=BLACK, bold=False, fill=None):
    c = ws[cell]; c.value = value
    c.font = Font(name=F, bold=bold, color=color, size=11)
    c.number_format = fmt if fmt else '#,##0'
    c.alignment = Alignment(horizontal='center', vertical='center')
    if fill: c.fill = fill
    return c

def header_row(ws, row, cols, texts):
    for col, txt in zip(cols, texts):
        c = ws[f'{col}{row}']; c.value = txt
        c.font = Font(name=F, bold=True, color='FFFFFF', size=11)
        c.fill = HEADFILL; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)

# ═════════════ Sheet 1: الفرضيات ═════════════
a = wb.active; a.title = 'الفرضيات'; style(a)
a.column_dimensions['A'].width = 42
for col in 'BCDEF': a.column_dimensions[col].width = 15

label(a, 'A1', 'الفرضيات — محرّكات النموذج المالي', bold=True, color=NAVY, size=15)
label(a, 'A2', 'الخلايا الزرقاء = مدخلات قابلة للتعديل  ·  الأسود = معادلات', color='777777', size=9)

label(a, 'A4', 'أولاً: محرّكات الإيراد', bold=True, color=GREEN, size=12, fill=SUBFILL)
a['B4'].fill = SUBFILL
label(a, 'A5', 'عدد البراندات — السنة 1'); num(a, 'B5', 50, '#,##0', BLUE, fill=YELLOW)
label(a, 'A6', 'متوسط الإنفاق السنوي للبراند (ريال)'); num(a, 'B6', 80000, MONEY, BLUE, fill=YELLOW)
label(a, 'A7', 'إجمالي قيمة المنشورات GMV — السنة 1'); num(a, 'B7', '=B5*B6')
label(a, 'A8', 'نمو GMV — السنة 2'); num(a, 'B8', 3.5, PCT, BLUE, fill=YELLOW)
label(a, 'A9', 'نمو GMV — السنة 3'); num(a, 'B9', 1.5, PCT, BLUE, fill=YELLOW)
label(a, 'A10', 'نمو GMV — السنة 4'); num(a, 'B10', 0.6667, PCT, BLUE, fill=YELLOW)
label(a, 'A11', 'نمو GMV — السنة 5'); num(a, 'B11', 0.58, PCT, BLUE, fill=YELLOW)

label(a, 'A13', 'ثانياً: العمولة (Take Rate)', bold=True, color=GREEN, size=12, fill=SUBFILL)
a['B13'].fill = SUBFILL
label(a, 'A14', 'نسبة الحملات العادية من الحجم'); num(a, 'B14', 0.70, PCT, BLUE, fill=YELLOW)
label(a, 'A15', 'عمولة الحملات العادية'); num(a, 'B15', 0.05, PCT, BLUE, fill=YELLOW)
label(a, 'A16', 'عمولة جلسات البث المباشر'); num(a, 'B16', 0.15, PCT, BLUE, fill=YELLOW)
label(a, 'A17', 'العمولة المجمّعة (المحسوبة)', bold=True); num(a, 'B17', '=B14*B15+(1-B14)*B16', PCT, GREEN, bold=True)
a['B17'].comment = Comment('= (نسبة الحملات×عمولتها) + (نسبة الجلسات×عمولتها) = 0.70×5% + 0.30×15% = 8%', 'WINوWIN')

label(a, 'A19', 'ثالثاً: الاستثمار والخصم', bold=True, color=GREEN, size=12, fill=SUBFILL)
a['B19'].fill = SUBFILL
label(a, 'A20', 'التكاليف التأسيسية (السنة 0)'); num(a, 'B20', 250000, MONEY, BLUE, fill=YELLOW)
label(a, 'A21', 'إجمالي الاستثمار المطلوب'); num(a, 'B21', 1200000, MONEY, BLUE, fill=YELLOW)
a['B21'].comment = Comment('يغطي: تأسيس 250 ألف + تمويل خسائر أول سنتين (~890 ألف) + احتياطي ~60 ألف', 'WINوWIN')
label(a, 'A22', 'معدل الخصم (WACC)'); num(a, 'B22', 0.12, PCT, BLUE, fill=YELLOW)

# Costs table
label(a, 'A24', 'رابعاً: المصاريف السنوية (ريال)', bold=True, color=GREEN, size=12, fill=SUBFILL)
for col in 'BCDEF': a[f'{col}24'].fill = SUBFILL
header_row(a, 25, ['A','B','C','D','E','F'], ['البند','السنة 1','السنة 2','السنة 3','السنة 4','السنة 5'])
costs = {
    26: ('الرواتب', [480000, 900000, 1500000, 2300000, 3400000]),
    27: ('التسويق', [300000, 500000, 700000, 900000, 1300000]),
    28: ('البنية التحتية والـ SMS', [60000, 180000, 350000, 550000, 900000]),
    29: ('مصاريف إدارية وقانونية', [80000, 150000, 250000, 450000, 680000]),
}
for row, (name, vals) in costs.items():
    label(a, f'A{row}', name); a[f'A{row}'].border = BORDER
    for i, col in enumerate('BCDEF'):
        num(a, f'{col}{row}', vals[i], MONEY, BLUE, fill=YELLOW); a[f'{col}{row}'].border = BORDER
label(a, 'A30', 'إجمالي المصاريف', bold=True, fill=TOTALFILL); a['A30'].border = BORDER
for col in 'BCDEF':
    num(a, f'{col}30', f'=SUM({col}26:{col}29)', MONEY, bold=True, fill=TOTALFILL); a[f'{col}30'].border = BORDER

# ═════════════ Sheet 2: النموذج ═════════════
m = wb.create_sheet('النموذج'); style(m)
m.column_dimensions['A'].width = 34
for col in 'BCDEF': m.column_dimensions[col].width = 16
label(m, 'A1', 'النموذج — قائمة الدخل المتوقعة (5 سنوات)', bold=True, color=NAVY, size=15)
label(m, 'A2', 'كل الأرقام بالريال السعودي، غير شاملة ضريبة القيمة المضافة (15%)', color='777777', size=9)
header_row(m, 3, ['A','B','C','D','E','F'], ['البند','السنة 1','السنة 2','السنة 3','السنة 4','السنة 5'])

# GMV
label(m, 'A4', 'إجمالي قيمة المنشورات (GMV)'); m['A4'].border = BORDER
num(m, 'B4', "=الفرضيات!B7");
num(m, 'C4', "=B4*(1+الفرضيات!B8)")
num(m, 'D4', "=C4*(1+الفرضيات!B9)")
num(m, 'E4', "=D4*(1+الفرضيات!B10)")
num(m, 'F4', "=E4*(1+الفرضيات!B11)")
# Revenue
label(m, 'A5', 'إيراد المنصة (عمولة)', bold=True); m['A5'].border = BORDER
for col in 'BCDEF':
    num(m, f'{col}5', f'={col}4*الفرضيات!$B$17', MONEY, GREEN, bold=True)
# Costs (link to assumptions row-by-row)
cost_map = {7:('الرواتب',26), 8:('التسويق',27), 9:('البنية والـ SMS',28), 10:('مصاريف إدارية وقانونية',29)}
for mrow,(name,arow) in cost_map.items():
    label(m, f'A{mrow}', name)
    for col in 'BCDEF':
        num(m, f'{col}{mrow}', f'=الفرضيات!{col}{arow}', MONEY, GREENTXT)
label(m, 'A11', 'إجمالي المصاريف', bold=True, fill=TOTALFILL)
for col in 'BCDEF':
    num(m, f'{col}11', f'=SUM({col}7:{col}10)', MONEY, bold=True, fill=TOTALFILL)
label(m, 'A13', 'صافي الربح / (الخسارة)', bold=True, color=NAVY)
for col in 'BCDEF':
    num(m, f'{col}13', f'={col}5-{col}11', MONEY, bold=True, fill=SUBFILL)
label(m, 'A14', 'الرصيد التراكمي', bold=True)
num(m, 'B14', '=B13')
for prev,col in zip('BCDE','CDEF'):
    num(m, f'{col}14', f'={prev}14+{col}13', MONEY, bold=True)
label(m, 'A16', 'هامش الربح الصافي')
for col in 'BCDEF':
    num(m, f'{col}16', f'=IF({col}5=0,0,{col}13/{col}5)', PCT)

# ═════════════ Sheet 3: التحليل ═════════════
an = wb.create_sheet('التحليل'); style(an)
an.column_dimensions['A'].width = 38
for col in 'BCDEFG': an.column_dimensions[col].width = 15
label(an, 'A1', 'التحليل المالي — الجدوى والعائد', bold=True, color=NAVY, size=15)
label(an, 'A3', 'التدفقات النقدية للمستثمر (السنة 0 = ضخ الاستثمار)', bold=True, color=GREEN, size=12, fill=SUBFILL)
for col in 'BCDEFG': an[f'{col}3'].fill = SUBFILL
header_row(an, 4, ['A','B','C','D','E','F','G'], ['البند','السنة 0','السنة 1','السنة 2','السنة 3','السنة 4','السنة 5'])
label(an, 'A5', 'صافي التدفق النقدي')
num(an, 'B5', '=-الفرضيات!B21')
num(an, 'C5', 0); num(an, 'D5', 0)   # Y1,Y2 losses funded by the raised investment
num(an, 'E5', '=النموذج!D13'); num(an, 'F5', '=النموذج!E13'); num(an, 'G5', '=النموذج!F13')
an['C5'].comment = Comment('خسائر السنتين 1-2 ممولة أصلاً من الاستثمار المضخوخ في السنة 0', 'WINوWIN')
label(an, 'A6', 'الرصيد التراكمي', bold=True)
num(an, 'B6', '=B5')
for prev,col in zip('BCDEF','CDEFG'):
    num(an, f'{col}6', f'={prev}6+{col}5', MONEY, bold=True)

label(an, 'A8', 'مؤشرات الجدوى', bold=True, color=GREEN, size=12, fill=SUBFILL)
an['B8'].fill = SUBFILL
label(an, 'A9', 'صافي القيمة الحالية NPV @ معدل الخصم'); num(an, 'B9', '=B5+NPV(الفرضيات!B22,C5:G5)', MONEY, NAVY, bold=True)
label(an, 'A10', 'معدل العائد الداخلي IRR'); num(an, 'B10', '=IFERROR(IRR(B5:G5),"غير محقق")', PCT, NAVY, bold=True)
label(an, 'A11', 'فترة الاسترداد (سنة)'); num(an, 'B11', '=3+(-E6)/F5', '0.0', NAVY, bold=True)
an['B11'].comment = Comment('التراكمي يتحول موجباً خلال السنة 4: 3 سنوات + (400,000 ÷ 1,800,000)', 'WINوWIN')
label(an, 'A12', 'هامش الربح الصافي — السنة 3'); num(an, 'B12', '=النموذج!D16', PCT, bold=True)
label(an, 'A13', 'هامش الربح الصافي — السنة 5'); num(an, 'B13', '=النموذج!F16', PCT, bold=True)
label(an, 'A14', 'نقطة التعادل التشغيلي'); label(an, 'B14', 'أوائل السنة 3', bold=True, align='center')

# ═════════════ Sheet 4: الحساسية ═════════════
se = wb.create_sheet('الحساسية'); style(se)
se.column_dimensions['A'].width = 34
for col in 'BCD': se.column_dimensions[col].width = 17
label(se, 'A1', 'تحليل الحساسية — 3 سيناريوهات', bold=True, color=NAVY, size=15)
label(se, 'A2', 'يُطبّق "معامل النمو" على معدلات نمو GMV. المصاريف ثابتة (افتراض متحفّظ).', color='777777', size=9)
header_row(se, 4, ['A','B','C','D'], ['المؤشر','متحفّظ','أساسي','متفائل'])
label(se, 'A5', 'معامل النمو');
num(se, 'B5', 0.7, '0.0x', BLUE, fill=YELLOW); num(se, 'C5', 1.0, '0.0x', BLUE, fill=YELLOW); num(se, 'D5', 1.3, '0.0x', BLUE, fill=YELLOW)

# GMV rows 6-10 (Y1..Y5) per scenario
label(se, 'A6', 'GMV — السنة 1')
for col in 'BCD': num(se, f'{col}6', '=الفرضيات!$B$7')
gyears = {7:('GMV — السنة 2','B8','6'), 8:('GMV — السنة 3','B9','7'), 9:('GMV — السنة 4','B10','8'), 10:('GMV — السنة 5','B11','9')}
for row,(name,grow,prev) in gyears.items():
    label(se, f'A{row}', name)
    for col in 'BCD':
        num(se, f'{col}{row}', f'={col}{prev}*(1+الفرضيات!${grow[0]}${grow[1:]}*{col}$5)')
# Revenue rows 12-16
label(se, 'A12', 'الإيراد — السنة 1')
for r_off,gr in zip(range(12,17), range(6,11)):
    label(se, f'A{r_off}', f'الإيراد — السنة {r_off-11}')
    for col in 'BCD':
        num(se, f'{col}{r_off}', f'={col}{gr}*الفرضيات!$B$17', MONEY, GREEN)
# Net rows 18-22 = Revenue - total cost (assumptions row30 B..F)
costcols = {18:'B',19:'C',20:'D',21:'E',22:'F'}  # year1..5 total cost columns in الفرضيات row30
for r,(ccol) in costcols.items():
    yr = r-17
    label(se, f'A{r}', f'صافي — السنة {yr}')
    for col in 'BCD':
        rev_row = 11+yr  # revenue rows 12..16
        num(se, f'{col}{r}', f'={col}{rev_row}-الفرضيات!${ccol}$30', MONEY, bold=True)

# Investor cash flow + NPV/IRR per scenario (rows 24-27)
label(se, 'A24', 'NPV @ معدل الخصم', bold=True, fill=TOTALFILL)
for col in 'BCD':
    # CF: t0=-invest, Y1=0, Y2=0, Y3..Y5 = net rows 20,21,22
    num(se, f'{col}24', f'=-الفرضيات!$B$21+NPV(الفرضيات!$B$22,0,0,{col}20,{col}21,{col}22)', MONEY, NAVY, bold=True, fill=TOTALFILL)
label(se, 'A25', 'معدل العائد الداخلي IRR', bold=True, fill=TOTALFILL)
for col in 'BCD':
    num(se, f'{col}25', f'=IFERROR(IRR({{-1}}*0+ (-الفرضيات!$B$21)),"—")', PCT, NAVY, bold=True, fill=TOTALFILL)

# IRR needs a contiguous range; build helper CF rows 27-32 per scenario then IRR over them
label(se, 'A27', 'التدفقات النقدية للمستثمر (للحساب):', color='777777', size=9)
cfmap = {28:('السنة 0', None), 29:('السنة 1', '0'), 30:('السنة 2','0'), 31:('السنة 3','20'), 32:('السنة 4','21'), 33:('السنة 5','22')}
for r,(name,src) in cfmap.items():
    label(se, f'A{r}', name, color='777777', size=9)
    for col in 'BCD':
        if src is None:
            num(se, f'{col}{r}', '=-الفرضيات!$B$21', MONEY, color='777777')
        elif src == '0':
            num(se, f'{col}{r}', 0, MONEY, color='777777')
        else:
            num(se, f'{col}{r}', f'={col}{src}', MONEY, color='777777')
# Fix IRR row25 to use the helper CF range
for col in 'BCD':
    se[f'{col}25'].value = f'=IFERROR(IRR({col}28:{col}33),"غير محقق")'

# ═════════════ Add 4th scenario (أسوأ حالة) to الحساسية ═════════════
se.column_dimensions['E'].width = 17
se['E4'].value = 'أسوأ حالة'; se['E4'].font = Font(name=F, bold=True, color='FFFFFF'); se['E4'].fill = HEADFILL
se['E4'].border = BORDER; se['E4'].alignment = Alignment(horizontal='center', vertical='center', readingOrder=2)
num(se, 'E5', 0.4, '0.0x', BLUE, fill=YELLOW)
num(se, 'E6', '=الفرضيات!$B$7')
for row,(name,grow,prev) in gyears.items():
    num(se, f'E{row}', f'=E{prev}*(1+الفرضيات!${grow[0]}${grow[1:]}*E$5)')
for r_off,gr in zip(range(12,17), range(6,11)):
    num(se, f'E{r_off}', f'=E{gr}*الفرضيات!$B$17', MONEY, GREEN)
for r,ccol in costcols.items():
    yr=r-17; num(se, f'E{r}', f'=E{11+yr}-الفرضيات!${ccol}$30', MONEY, bold=True)
num(se, 'E24', '=-الفرضيات!$B$21+NPV(الفرضيات!$B$22,0,0,E20,E21,E22)', MONEY, NAVY, bold=True, fill=TOTALFILL)
for r,(name,src) in cfmap.items():
    if src is None: num(se, f'E{r}', '=-الفرضيات!$B$21', MONEY, color='777777')
    elif src=='0': num(se, f'E{r}', 0, MONEY, color='777777')
    else: num(se, f'E{r}', f'=E{src}', MONEY, color='777777')
num(se, 'E25', '=IFERROR(IRR(E28:E33),"غير محقق")', PCT, NAVY, bold=True, fill=TOTALFILL)

# ═════════════ Sheet 5: الاقتصاد الوحدوي (Unit Economics) ═════════════
ue = wb.create_sheet('الاقتصاد الوحدوي'); style(ue)
ue.column_dimensions['A'].width = 42
for col in 'BC': ue.column_dimensions[col].width = 18
label(ue, 'A1', 'الاقتصاد الوحدوي (Unit Economics)', bold=True, color=NAVY, size=15)
label(ue, 'A2', 'يوضّح ربحية اكتساب كل مستخدم وكل براند — أساس قابلية التوسّع', color='777777', size=9)
header_row(ue, 4, ['A','B','C'], ['المؤشر', 'المستخدم', 'البراند'])
label(ue, 'A5', 'عدد الوحدات — السنة 1'); num(ue,'B5',20000,'#,##0',BLUE,fill=YELLOW); num(ue,'C5','=الفرضيات!B5',None,GREENTXT)
label(ue, 'A6', 'تكلفة الاكتساب CAC (ريال)'); num(ue,'B6',20,MONEY,BLUE,fill=YELLOW); num(ue,'C6',1000,MONEY,BLUE,fill=YELLOW)
ue['B6'].comment = Comment('نمو عضوي + إحالة → تكلفة منخفضة', 'WINوWIN'); ue['C6'].comment=Comment('جهد مبيعات مباشر','WINوWIN')
label(ue, 'A7', 'إيراد المنصة لكل وحدة سنوياً (ريال)')
num(ue,'B7','=النموذج!B5/B5', MONEY, GREEN)   # revenue Y1 / users
num(ue,'C7','=الفرضيات!B6*الفرضيات!B17', MONEY, GREEN)  # brand spend * take
label(ue, 'A8', 'متوسط بقاء الوحدة (سنوات)'); num(ue,'B8',2.5,'0.0',BLUE,fill=YELLOW); num(ue,'C8',3.0,'0.0',BLUE,fill=YELLOW)
label(ue, 'A9', 'هامش المساهمة'); num(ue,'B9',0.75,PCT,BLUE,fill=YELLOW); num(ue,'C9',0.75,PCT,BLUE,fill=YELLOW)
label(ue, 'A10', 'القيمة الدائمة LTV (ريال)', bold=True)
num(ue,'B10','=B7*B8*B9', MONEY, NAVY, bold=True, fill=SUBFILL)
num(ue,'C10','=C7*C8*C9', MONEY, NAVY, bold=True, fill=SUBFILL)
label(ue, 'A11', 'نسبة LTV/CAC', bold=True)
num(ue,'B11','=B10/B6', MULT, NAVY, bold=True); num(ue,'C11','=C10/C6', MULT, NAVY, bold=True)
ue['B12'] = None
label(ue, 'A13', 'قاعدة إبهام: LTV/CAC ≥ 3x تعني اكتساباً صحياً وقابلاً للتوسّع.', color='777777', size=9)
label(ue, 'A14', 'البراندات هي محرّك الربح الأساسي؛ المستخدمون يُكتسبون بتكلفة منخفضة لتغذية الشبكة.', color=NAVY, size=10, bold=True)

# ═════════════ Sheet 6: التوقعات الشهرية (24 شهراً) ═════════════
mo = wb.create_sheet('التوقعات الشهرية'); style(mo)
mo.column_dimensions['A'].width = 12
for col in 'BCDEF': mo.column_dimensions[col].width = 16
label(mo, 'A1', 'التوقعات الشهرية — أول 24 شهراً', bold=True, color=NAVY, size=15)
label(mo, 'A2', 'مبنية على منحنى تصاعدي يوازي إجمالي السنة؛ الرصيد الافتتاحي = الاستثمار − التأسيس', color='777777', size=9)
# opening cash cell
label(mo, 'A3', 'الرصيد النقدي الافتتاحي'); num(mo,'B3','=الفرضيات!B21-الفرضيات!B20', MONEY, NAVY, bold=True, fill=YELLOW)
header_row(mo, 5, ['A','B','C','D','E','F'], ['الشهر','GMV','الإيراد','المصاريف','الصافي','الرصيد النقدي'])
# months 1..24 as rows 6..29. weight ramp: m/78 within each year
for i in range(1,25):
    r = 5+i
    label(mo, f'A{r}', f'شهر {i}', align='center')
    if i<=12:
        gmv_f = f'=النموذج!$B$4*({i}/78)'; opex_f = '=الفرضيات!$B$30/12'
    else:
        gmv_f = f'=النموذج!$C$4*({i-12}/78)'; opex_f = '=الفرضيات!$C$30/12'
    num(mo, f'B{r}', gmv_f)
    num(mo, f'C{r}', f'=B{r}*الفرضيات!$B$17', MONEY, GREEN)
    num(mo, f'D{r}', opex_f)
    num(mo, f'E{r}', f'=C{r}-D{r}', MONEY, bold=True)
    prev = 'B3' if i==1 else f'F{r-1}'
    num(mo, f'F{r}', f'={prev}+E{r}', MONEY, NAVY, bold=True)
# summary
label(mo, 'A31', 'أدنى رصيد نقدي خلال 24 شهراً'); num(mo,'B31','=MIN(F6:F29)', MONEY, NAVY, bold=True)
label(mo, 'A32', 'متوسط الحرق الشهري — السنة 1 (Burn)'); num(mo,'B32','=AVERAGE(E6:E17)', '#,##0;(#,##0)', BLACK, True)
label(mo, 'A33', 'الرصيد النقدي بنهاية الشهر 24'); num(mo,'B33','=F29', MONEY, NAVY, bold=True)
label(mo, 'A34', 'ملاحظة: بقاء أدنى رصيد موجباً = كفاية الاستثمار حتى الوصول للربحية.', color='777777', size=9)

# ═════════════ Sheet 7: التمويل والعائد ═════════════
fu = wb.create_sheet('التمويل والعائد'); style(fu)
fu.column_dimensions['A'].width = 40
for col in 'BC': fu.column_dimensions[col].width = 18
label(fu, 'A1', 'خطة استخدام التمويل والعائد للمستثمر', bold=True, color=NAVY, size=15)
label(fu, 'A2', 'أرقام التقييم والعائد استرشادية وقابلة للتفاوض', color='777777', size=9)
label(fu, 'A4', 'أولاً: استخدام التمويل (1.2 مليون ريال)', bold=True, color=GREEN, size=12, fill=SUBFILL); fu['B4'].fill=SUBFILL
header_row(fu, 5, ['A','B','C'], ['البند','النسبة','المبلغ (ريال)'])
uses = [('التسويق والنمو',0.40),('الرواتب والفريق',0.40),('التقنية والبنية والـ SMS',0.08),('قانوني وتراخيص وإداري',0.10),('احتياطي نقدي',0.02)]
r=6
for name,pct in uses:
    label(fu,f'A{r}',name); num(fu,f'B{r}',pct,PCT,BLUE,fill=YELLOW); num(fu,f'C{r}',f'=B{r}*الفرضيات!$B$21', MONEY); r+=1
label(fu,f'A{r}','الإجمالي', bold=True, fill=TOTALFILL); num(fu,f'B{r}',f'=SUM(B6:B{r-1})',PCT,bold=True,fill=TOTALFILL); num(fu,f'C{r}',f'=SUM(C6:C{r-1})',MONEY,bold=True,fill=TOTALFILL)

label(fu, 'A13', 'ثانياً: شروط الاستثمار (استرشادي)', bold=True, color=GREEN, size=12, fill=SUBFILL); fu['B13'].fill=SUBFILL
label(fu,'A14','مبلغ الجولة'); num(fu,'B14','=الفرضيات!B21', MONEY, GREENTXT)
label(fu,'A15','حصة المستثمر'); num(fu,'B15',0.22,PCT,BLUE,fill=YELLOW)
label(fu,'A16','التقييم بعد الجولة (Post-money)'); num(fu,'B16','=B14/B15', MONEY, NAVY, bold=True)
label(fu,'A17','التقييم قبل الجولة (Pre-money)'); num(fu,'B17','=B16-B14', MONEY, NAVY, bold=True)
label(fu,'A18','حصة المؤسسين'); num(fu,'B18','=1-B15',PCT,bold=True)

label(fu, 'A20', 'ثالثاً: تقدير العائد عند الخروج (السنة 5)', bold=True, color=GREEN, size=12, fill=SUBFILL); fu['B20'].fill=SUBFILL
label(fu,'A21','مضاعف التقييم على الإيراد (Revenue Multiple)'); num(fu,'B21',5,MULT,BLUE,fill=YELLOW)
label(fu,'A22','إيراد السنة 5'); num(fu,'B22','=النموذج!F5', MONEY, GREENTXT)
label(fu,'A23','قيمة الشركة عند الخروج'); num(fu,'B23','=B21*B22', MONEY, NAVY, bold=True)
label(fu,'A24','حصة المستثمر عند الخروج'); num(fu,'B24','=B23*B15', MONEY, NAVY, bold=True)
label(fu,'A25','مضاعف رأس المال MOIC'); num(fu,'B25','=B24/B14', MULT, NAVY, bold=True)
label(fu,'A26','معدل العائد السنوي التقريبي على 5 سنوات'); num(fu,'B26','=(B24/B14)^(1/5)-1', PCT, NAVY, bold=True)
label(fu, 'A28', 'خطة الخروج: استحواذ استراتيجي (مجموعات إعلام/تسويق أو منصة إقليمية)، أو بيع حصص، أو إدراج مستقبلي في السوق الموازية (نمو). الأفق 5–7 سنوات.', color=NAVY, size=10, bold=True)

# Force Excel to recalculate all formulas when the file is opened
try:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcId = 0
except Exception:
    pass

OUT = os.path.join(os.path.dirname(__file__), 'WinWin-Financial-Model.xlsx')
wb.save(OUT)
print('SAVED WinWin-Financial-Model.xlsx')

# ── Independent verification of the base scenario (Python) ──
brands, spend = 50, 80000
gmv1 = brands*spend
g = [3.5,1.5,0.6667,0.58]
gmv=[gmv1]
for gr in g: gmv.append(gmv[-1]*(1+gr))
take=0.08
rev=[x*take for x in gmv]
costs_tot=[920000,1730000,2800000,4200000,6280000]
net=[rev[i]-costs_tot[i] for i in range(5)]
print('GMV(M):', [round(x/1e6,2) for x in gmv])
print('Revenue(M):', [round(x/1e6,3) for x in rev])
print('Net(k):', [round(x/1e3) for x in net])
# investor CF
cf=[-1200000,0,0,net[2],net[3],net[4]]
def npv(rate,flows): return sum(f/(1+rate)**i for i,f in enumerate(flows))
def irr(flows):
    lo,hi=-0.9,5.0
    for _ in range(200):
        mid=(lo+hi)/2
        if npv(mid,flows)>0: lo=mid
        else: hi=mid
    return mid
print('NPV@12% (investor):', round(npv(0.12,cf)))
print('IRR (investor):', round(irr(cf)*100,1),'%')
print('cumulative(k):', [round(sum(cf[:i+1])/1e3) for i in range(len(cf))])
print('payback:', round(3+(-sum(cf[:4]))/cf[4],2),'yr' if cf[4] else '')
